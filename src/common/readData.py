import csv
import os
from openpyxl import load_workbook
import json
from src.common.getPath import GetPath


class ReadCsvData:
    def read_csv_test_data(self,line,fileName):
        # 1）先找到这个csv文件 （涉及到：路径+文件名.csv）
        dataFilePath = os.path.join(GetPath().get_data_path(),fileName)
        # 2）打开这个csv文件
        with open(dataFilePath,'r',encoding='utf8') as f:
            # 3） 读取csv文件中的内容。 （csv库：因为是csv文件；）
            dataContent = csv.reader(f)
            # 4） 把读取到的内容转换成python的列表类型，方便后续的处理
            listData = list(dataContent)
        # 5） 把读取到的数据返回出去（按行返回，因为一行对应一个用例） 。（return ）
        return listData[line]

    def read_csv_config_data(self, fileName):
        # 1） 先找到这个csv文件。 （涉及到：路径 + 文件名.csv   os模块）
        currentPath = os.path.dirname(__file__)
        configFilePath = os.path.join(GetPath().get_config_path(), fileName)
        # 2） 打开这个csv文件。 （with  open）
        with open(configFilePath, 'r', encoding='utf8') as f:
            # 3） 读取csv文件中的内容。 （csv库：因为是csv文件；）
            configContent = csv.reader(f)
            # 4） 把读取到的内容转换成python的字典类型，方便后续的处理
            dictData = dict(configContent)
        # 5） 把读取到的数据返回出去。（return ）
        return dictData

class WriteToCsvData:
    # 通过csv.write方法写入csv文件，需要传入列表套元组或者列表套列表 data = [(,,),(,,)]
    def write_to_csv_file(self, fileName, data):
        csvFilePath=os.path.join(GetPath().get_config_path(),fileName)
        # 打开csv文件，用w方法
        with open(csvFilePath, 'w', encoding='utf8') as f:
            # 循环写入文件
            for line in data:
                csv.writer(f).writerow(line)

    # 通过csv.DictWriter方法写入csv文件，可以先写入字段，以列表格式fieldnames =[],再写值，需要传入列表套字典 data = [{},{}]
    def dictwriteto_csv_file(self, fileName, fieldnames, data):
        csvFilePath = os.path.join(GetPath().get_config_path(),fileName)
        # 打开csv文件，用w方法
        with open(csvFilePath, 'w', encoding='utf8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            # 将字段写入首行
            writer.writeheader()
            # 循环写入文件
            for line in data:
                writer.writerow(line)


# 定义一个类，实现获取到工作簿、工作表、单元格中的值、工作表的最大行号
class ReadExcelData():
    # 定义一个构造方法，用来获取到工作簿
    def __init__(self, fileName):
        """
        :param fileName:表示要操作的excel文件名。
        """
        try:
            # 获取到工作簿
            self.workbook = load_workbook(os.path.join(GetPath().get_data_path(), fileName))
        except:  # No such file or directory
            raise Exception("File not exit, please check")

    # 定义一个方法，用来获取到工作簿中的工作表
    def get_worksheet_by_name(self, sheetName):
        """
        :param sheetName:表示工作表的名称。
        """
        try:
            worksheet = self.workbook[sheetName]
            return worksheet
        except Exception as e:
            raise e

    # 定义一个方法，用来获取到某个工作表中最大的行号
    def get_max_rows(self, sheetName):
        try:
            # 返回获取到某个工作表中的最大的行号。
            return self.get_worksheet_by_name(sheetName).max_row  # 调用max_row属性获取到最大的行号
        except Exception as e:
            raise e

    # 定义一个方法，用来获取到某个工作表中的某个单元格中的值
    # 获取单元格中的值有两种方法：1、通过坐标来获取（例如：A3，E3等）   2、通过行和列的方式来获取
    def get_value_of_cell(self, sheetName, coordinate=None, rowNum=None,
                          columnNum=None):
        # 如果传入的坐标不为空，才可以获取到值  ====> 1、通过坐标来获取（例如：A3，E3等）
        if coordinate != None:
            try:
                cell = self.get_worksheet_by_name(sheetName)[coordinate]  # 表示唯一的确定某个单元格
                return cell.value
            except Exception as e:
                raise e
        # 如果传入的坐标为空，但是传入的行和列不为空，则也可以读取到值  ====》 2、通过行和列的方式来获取
        elif coordinate == None and rowNum != None and columnNum != None:
            try:
                return self.get_worksheet_by_name(sheetName).cell(row=rowNum, column=columnNum).value
            except Exception as e:
                raise e
        # 如果坐标为空，或者 行和列有一个为空，两个都为空，则无法读取到值
        else:
            raise Exception("No coordinate specified, please check")

    # 定义一个函数，用来从excel表格中读取到数据，保存成一个列表套列表或者列表套元组的格式
    def read_excel_data(self, sheetName, columnNum):
        # 1、遍历某个工作表中的每一行，获取数据。
        excelData = []
        # 获取到最大化的行号
        maxRows = self.get_max_rows(sheetName)
        for row in range(2, maxRows + 1):
            # 获取到请求参数列的值
            excelData.append(self.get_value_of_cell(sheetName, rowNum=row, columnNum=columnNum))
            # 获取到返回结果列的值，并需要把读取到的值由字符串转换成字典类型（json.loads() 作用：把字符串转换成字典）
            # expected_data = json.loads(self.get_value_of_cell(sheetName, rowNum=row, columnNum=11))
        # 2、返回获取到所有的数据
        return excelData  # 一定要写在for循环外

